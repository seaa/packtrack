from django.contrib import admin
from django.http import HttpResponse
from django.utils import timezone
from import_export import resources, fields
from import_export.admin import ImportExportModelAdmin, ImportExportActionModelAdmin
from import_export.widgets import ForeignKeyWidget, ManyToManyWidget
from .models import *
from utils.models import *


class OriginItemInline(admin.StackedInline):
    model = Origin
    extra = 1


class DestinationItemInline(admin.StackedInline):
    model = Destination
    extra = 1


class PackageItemInline(admin.TabularInline):
    model = OrderPackage
    extra = 1


class OrderStatusInline(admin.TabularInline):
    model = OrderStatus
    readonly_fields = ['st_update']
    extra = 1

    def get_status(self, obj):
        return obj.get_status_display()


class OrderAdmin(admin.ModelAdmin):

    model = Order
    fieldsets = (('TITULO', {
                    'fields': ('id', 'title', 'description', 'ord_price')
                    }),
                 ('CLIENTE', {
                     'fields': (('client', 'request_id'),)
                 }),
                 ('TIEMPO', {
                     'fields': ('created_at', 'start_time', 'end_time', 'duration')
                 }),
                 )
    readonly_fields = ['id', 'created_at']

    # list of fields to display in django admin
    list_display = ('title', 'id', 'client', 'created_at')

    # if you want django admin to show the search bar, just add this line
    # search_fields = ('client','created_at')
    filter_horizontal = ()
    list_filter = ('client', 'created_at')

    # inlines
    inlines = (OriginItemInline, DestinationItemInline, PackageItemInline, OrderStatusInline,)

def export_xlsx(modeladmin, request, queryset):

        import openpyxl
        import re
        from openpyxl.utils import get_column_letter
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = re.sub(' ', '_', '{}-export.xlsx'.format(timezone.localtime()))
        response['Content-Disposition'] = 'attachment; filename={}'.format(filename)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "MyModel"

        row_num = 0

        columns = [
            (u"Task Title", 30),
            (u"Start Date", 20),
            (u"End Date", 20),
            (u"Address", 50),
            (u"Latitude", 20),
            (u"Longitude", 20),
            (u"Description", 30),
            (u"Order ID", 5),
            (u"Performer", 5),
            (u"Price", 20),
            (u"Client", 20),
            (u"Req_id", 20),
        ]

        for col_num in range(len(columns)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = columns[col_num][0]
            # c.style.font.bold = True
            # set column width
            ws.column_dimensions[get_column_letter(col_num + 1)].width = columns[col_num][1]

        for obj in queryset:
            dest = Destination.objects.filter(order=obj.pk).values()[0]
            status = OrderStatus.objects.filter(order=obj.pk).last()
            wk = ('street', 'house_num', 'suburb', 'city', 'province', 'country', 'pos_code')
            address = re.sub(',', '', ', '.join(str(value) for value in dict(zip(wk, [dest[k] for k in wk])).values() if value), 1)
            row_num += 1
            row = [
                dest['name'],
                obj.start_time,
                obj.end_time,
                address,
                dest['latitude'],
                dest['longitude'],
                obj.description,
                obj.pk,
                status.provider.prov_name,
                obj.ord_price,
                obj.client.client_name,
                obj.request_id,
            ]
            for col_num in range(len(row)):
                c = ws.cell(row=row_num + 1, column=col_num + 1)
                c.value = row[col_num]
                # c.style.alignment.wrap_text = True

        wb.save(response)

        return response

    actions = [export_xlsx]


admin.site.register(Order, OrderAdmin)

admin.site.site_header = 'PackGO Admin'
admin.site.site_title = 'PackGO Logistica'